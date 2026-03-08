from flask import Flask, render_template, request, redirect, url_for, session, flash, send_from_directory, send_file
from authlib.integrations.flask_client import OAuth
from flask_mail import Mail, Message
from models import db, Faculty, Achievements, Admin
import os
from werkzeug.utils import secure_filename
from werkzeug.middleware.proxy_fix import ProxyFix
import datetime
import json
import csv
import io
from openpyxl import Workbook
from fpdf import FPDF
import uuid
import socket
from dotenv import load_dotenv

load_dotenv()

GOOGLE_CLIENT_ID = os.environ.get("GOOGLE_CLIENT_ID")
GOOGLE_CLIENT_SECRET = os.environ.get("GOOGLE_CLIENT_SECRET")

app = Flask(__name__)
app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)
database_url = os.getenv("DATABASE_URL")

if database_url:
    # Render provides postgres:// but SQLAlchemy requires postgresql://
    database_url = database_url.replace("postgres://", "postgresql://", 1)
    app.config["SQLALCHEMY_DATABASE_URI"] = database_url
else:
    app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///database.db"
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
UPLOAD_FOLDER = os.path.join(os.getcwd(), "uploads")

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.secret_key = 'your_secret_key_here'  # Change this to a random secret key
app.config["MAIL_SERVER"] = os.environ.get("MAIL_SERVER", "smtp.gmail.com")
app.config["MAIL_PORT"] = int(os.environ.get("MAIL_PORT", 587))
app.config["MAIL_USE_TLS"] = os.environ.get("MAIL_USE_TLS", "true").lower() == "true"
app.config["MAIL_USE_SSL"] = os.environ.get("MAIL_USE_SSL", "false").lower() == "true"
app.config["MAIL_USERNAME"] = os.environ.get("MAIL_USERNAME")
app.config["MAIL_PASSWORD"] = os.environ.get("MAIL_PASSWORD")
app.config["MAIL_DEFAULT_SENDER"] = os.environ.get("MAIL_DEFAULT_SENDER", app.config["MAIL_USERNAME"])
app.config["MAIL_SUPPRESS_SEND"] = os.environ.get("MAIL_SUPPRESS_SEND", "false").lower() == "true"

# Google OAuth configuration with Authlib
oauth = OAuth(app)
google = oauth.register(
    name='google',
    client_id=GOOGLE_CLIENT_ID,
    client_secret=GOOGLE_CLIENT_SECRET,
    server_metadata_url='https://accounts.google.com/.well-known/openid-configuration',
    client_kwargs={'scope': 'openid email profile'}
)

db.init_app(app)
mail = Mail(app)

# File upload configuration
MAX_FILE_SIZE = 20 * 1024 * 1024  # 20MB
ALLOWED_CERTIFICATE_EXTENSIONS = {'pdf'}
ALLOWED_PHOTO_EXTENSIONS = {'jpg', 'jpeg', 'png'}

# Create upload folders if they don't exist
UPLOAD_SUBFOLDERS = [
    app.config["UPLOAD_FOLDER"],
    os.path.join(app.config["UPLOAD_FOLDER"], "certificates"),
    'static/uploads',
    'static/uploads/profile_photos'
]
for folder in UPLOAD_SUBFOLDERS:
    if not os.path.exists(folder):
        os.makedirs(folder)
        print(f"Created folder: {folder}")

def save_upload_file(file, upload_type):
    """Save uploaded file with validation and unique filename."""
    if not file or file.filename == '':
        return None, 'No file selected'
    
    filename = secure_filename(file.filename)
    file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
    
    # Check file extension
    if upload_type == 'certificate' and file_ext not in ALLOWED_CERTIFICATE_EXTENSIONS:
        return None, 'Only PDF files are allowed.'
    elif upload_type == 'photo' and file_ext not in ALLOWED_PHOTO_EXTENSIONS:
        return None, 'Only JPG, JPEG, and PNG files are allowed.'
    
    # Check file size
    file.seek(0, os.SEEK_END)
    file_size = file.tell()
    file.seek(0)
    if file_size > MAX_FILE_SIZE:
        return None, 'File size exceeds 20MB limit.'
    
    # Generate unique filename with uuid
    unique_filename = f"{uuid.uuid4()}_{filename}"
    
    if upload_type == 'certificate':
        file_path = os.path.join(app.config["UPLOAD_FOLDER"], "certificates", unique_filename)
        relative_path = file_path
    elif upload_type == 'photo':
        file_path = os.path.join('static/uploads/profile_photos', unique_filename)
        relative_path = f"uploads/profile_photos/{unique_filename}"
    else:
        return None, 'Invalid upload type'
    
    file.save(file_path)
    return relative_path, None

def send_achievement_submission_email(faculty, achievement):
    admin_email = os.environ.get("ADMIN_EMAIL")
    if not admin_email:
        admin = Admin.query.first()
        admin_email = admin.email if admin else None

    if not admin_email or not app.config.get("MAIL_USERNAME") or not app.config.get("MAIL_PASSWORD"):
        return

    msg = Message(
        subject="New Achievement Submitted",
        recipients=[admin_email]
    )
    msg.body = (
        f"Faculty Name: {faculty.name}\n"
        f"Achievement Type: {achievement.type}\n"
        f"Title: {achievement.title}"
    )
    mail.send(msg)

def init_db():
    """Initialize the database by creating all tables."""
    with app.app_context():
        db.create_all()
        print("Database initialized and tables created.")
        
        # Check and create test faculty user
        faculty = Faculty.query.filter_by(email='test@vnrvjiet.in').first()
        if not faculty:
            new_faculty = Faculty(
                name='Test Faculty',
                email='test@vnrvjiet.in',
                password='1234',
                department='CSE',
                designation='Professor'
            )
            db.session.add(new_faculty)
            db.session.commit()
            print("Test faculty user created.")
        
        # Ensure default admin exists
        admin = Admin.query.filter_by(email='admin@vnrvjiet.in').first()
        if not admin:
            new_admin = Admin(
                name='Super Admin',
                email='admin@vnrvjiet.in',
                password='admin123'
            )
            db.session.add(new_admin)
            db.session.commit()
            print("Default admin user created.")

@app.route('/')
def home():
    return render_template('home.html')

@app.route('/faculty/login', methods=['GET', 'POST'])
def faculty_login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        
        # Check email domain
        if not email.lower().endswith('@vnrvjiet.in'):
            flash('Only VNRVJIET faculty email accounts are allowed.')
            return redirect(url_for('faculty_login'))
        
        # Check if faculty exists
        faculty = Faculty.query.filter_by(email=email).first()
        if not faculty:
            flash('Faculty account not registered. Please register first.')
            return redirect(url_for('faculty_login'))
        
        # Check password
        if faculty.password != password:
            flash('Invalid password')
            return redirect(url_for('faculty_login'))
        
        session['faculty_id'] = faculty.faculty_id
        session['user_type'] = 'faculty'
        return redirect(url_for('faculty_dashboard'))
    return render_template('login.html')

@app.route('/login/google')
def login_google():
    redirect_uri = url_for('google_callback', _external=True, _scheme='https')
    return oauth.google.authorize_redirect(redirect_uri)

@app.route('/google/callback')
def google_callback():
    token = google.authorize_access_token()
    if not token:
        flash('Failed to authorize with Google.')
        return redirect(url_for('faculty_login'))
    
    user_info = token.get('userinfo')
    if not user_info:
        flash('Failed to retrieve user information from Google.')
        return redirect(url_for('faculty_login'))
    
    email = user_info.get('email', '')
    if not email.endswith("@vnrvjiet.in"):
        flash("Only VNRVJIET faculty are allowed to login.")
        return redirect(url_for("faculty_login"))
    
    faculty = Faculty.query.filter_by(email=email).first()
    if not faculty:
        flash('Faculty account not registered. Please register first.')
        return redirect(url_for('faculty_login'))
    
    session['faculty_id'] = faculty.faculty_id
    session['user_type'] = 'faculty'
    return redirect(url_for('faculty_dashboard'))

@app.route('/faculty/register', methods=['GET', 'POST'])
def faculty_register():
    if request.method == 'POST':
        name = request.form['name']
        email = request.form['email']
        password = request.form['password']
        department = request.form['department']
        designation = request.form['designation']
        
        # validate email domain
        if not email.lower().endswith('@vnrvjiet.in'):
            flash('Only VNRVJIET faculty email addresses (@vnrvjiet.in) are allowed.')
            return redirect(url_for('faculty_register'))
        
        if Faculty.query.filter_by(email=email).first():
            flash('Email already registered')
            return redirect(url_for('faculty_register'))
        
        # Handle profile photo upload (optional)
        profile_photo = None
        profile_photo_file = request.files.get('profile_photo')
        if profile_photo_file and profile_photo_file.filename != '':
            profile_photo, error = save_upload_file(profile_photo_file, 'photo')
            if error:
                flash(error)
                return redirect(url_for('faculty_register'))
        
        new_faculty = Faculty(name=name, email=email, password=password, department=department, designation=designation, profile_photo=profile_photo)
        db.session.add(new_faculty)
        db.session.commit()
        flash('Registration successful! Please login.')
        return redirect(url_for('faculty_login'))
    return render_template('faculty_register.html')

@app.route('/faculty/dashboard')
def faculty_dashboard():
    if 'faculty_id' not in session or session.get('user_type') != 'faculty':
        return redirect(url_for('faculty_login'))
    faculty = Faculty.query.get(session['faculty_id'])
    approved_achievements = [a for a in faculty.achievements if (a.status or '').lower() == 'approved']
    return render_template('dashboard.html', faculty=faculty, approved_achievements=approved_achievements)

@app.route('/faculty/logout')
def faculty_logout():
    session.pop('faculty_id', None)
    session.pop('user_type', None)
    return redirect(url_for('home'))

@app.route('/faculty/add_achievement', methods=['GET', 'POST'])
def add_achievement():
    if 'faculty_id' not in session or session.get('user_type') != 'faculty':
        return redirect(url_for('faculty_login'))
    if request.method == 'POST':
        title = request.form['title']
        achievement_type = request.form['type']
        description = request.form['description']
        date_str = request.form['date']
        date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
        
        # Handle proof file
        proof_file = request.files['proof_file']
        if proof_file and proof_file.filename.endswith('.pdf'):
            filename = secure_filename(proof_file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            proof_file.save(file_path)
        else:
            flash('Please upload a valid PDF file for proof.')
            return redirect(url_for('add_achievement'))
        
        # Handle certificate file (optional)
        certificate_file = request.files.get('certificate')
        certificate_path = None
        if certificate_file and certificate_file.filename != '':
            certificate_path, error = save_upload_file(certificate_file, 'certificate')
            if error:
                flash(error)
                return redirect(url_for('add_achievement'))
        
        achievement = Achievements(
            faculty_id=session['faculty_id'],
            title=title,
            type=achievement_type,
            description=description,
            date=date,
            proof_file=file_path,
            certificate_file=certificate_path,
            status='pending'
        )
        db.session.add(achievement)
        db.session.commit()
        try:
            faculty = Faculty.query.get(session['faculty_id'])
            send_achievement_submission_email(faculty, achievement)
        except Exception as e:
            print(f"Email notification failed: {e}")
        flash('Achievement added successfully!')
        return redirect(url_for('faculty_dashboard'))
    return render_template('add_achievement.html')

@app.route('/faculty/view_achievements')
def view_achievements():
    if 'faculty_id' not in session or session.get('user_type') != 'faculty':
        return redirect(url_for('faculty_login'))
    achievements = Achievements.query.filter_by(faculty_id=session['faculty_id']).all()
    return render_template('view_achievements.html', achievements=achievements)

@app.route('/faculty/profile')
def faculty_profile():
    if 'faculty_id' not in session or session.get('user_type') != 'faculty':
        return redirect(url_for('faculty_login'))
    faculty = Faculty.query.get(session['faculty_id'])
    achievements = Achievements.query.filter_by(faculty_id=session['faculty_id']).all()
    total_achievements = len(achievements)
    publications = len([a for a in achievements if a.type == 'Publication'])
    workshops = len([a for a in achievements if a.type == 'Workshop'])
    patents = len([a for a in achievements if a.type == 'Patent'])
    return render_template('faculty_profile.html', faculty=faculty, achievements=achievements, total_achievements=total_achievements, publications=publications, workshops=workshops, patents=patents)

@app.route('/faculty/edit_profile', methods=['GET', 'POST'])
def edit_profile():
    if 'faculty_id' not in session or session.get('user_type') != 'faculty':
        return redirect(url_for('faculty_login'))
    faculty = Faculty.query.get(session['faculty_id'])
    if request.method == 'POST':
        faculty.name = request.form['name']
        faculty.department = request.form['department']
        faculty.designation = request.form['designation']
        new_password = request.form.get('password')
        if new_password:
            faculty.password = new_password
        
        # Handle profile photo upload (optional)
        profile_photo_file = request.files.get('profile_photo')
        if profile_photo_file and profile_photo_file.filename != '':
            profile_photo, error = save_upload_file(profile_photo_file, 'photo')
            if error:
                flash(error)
                return redirect(url_for('edit_profile'))
            faculty.profile_photo = profile_photo
        
        db.session.commit()
        flash('Profile updated successfully!')
        return redirect(url_for('faculty_profile'))
    return render_template('edit_profile.html', faculty=faculty)

@app.route('/faculty/delete_achievement/<int:achievement_id>', methods=['POST'])
def delete_achievement(achievement_id):
    if 'faculty_id' not in session or session.get('user_type') != 'faculty':
        return redirect(url_for('faculty_login'))
    achievement = Achievements.query.get_or_404(achievement_id)
    if achievement.faculty_id != session['faculty_id']:
        flash('Unauthorized access')
        return redirect(url_for('view_achievements'))
    # Optionally delete the file
    if achievement.proof_file and os.path.exists(achievement.proof_file):
        os.remove(achievement.proof_file)
    db.session.delete(achievement)
    db.session.commit()
    flash('Achievement deleted successfully!')
    return redirect(url_for('view_achievements'))

@app.route('/admin/approve_achievement/<int:achievement_id>', methods=['POST'])
def approve_achievement(achievement_id):
    if 'admin_id' not in session or session.get('user_type') != 'admin':
        return redirect(url_for('admin_login'))
    achievement = Achievements.query.get_or_404(achievement_id)
    achievement.status = 'approved'
    db.session.commit()
    flash('Achievement approved!')
    return redirect(request.referrer or url_for('admin_dashboard'))

@app.route('/admin/reject_achievement/<int:achievement_id>', methods=['POST'])
def reject_achievement(achievement_id):
    if 'admin_id' not in session or session.get('user_type') != 'admin':
        return redirect(url_for('admin_login'))
    achievement = Achievements.query.get_or_404(achievement_id)
    achievement.status = 'rejected'
    db.session.commit()
    flash('Achievement rejected!')
    return redirect(request.referrer or url_for('admin_dashboard'))

@app.route('/admin/download_certificate/<int:achievement_id>')
def download_certificate(achievement_id):
    if 'admin_id' not in session or session.get('user_type') != 'admin':
        return redirect(url_for('admin_login'))
    achievement = Achievements.query.get_or_404(achievement_id)
    if not achievement.certificate_file or not os.path.exists(achievement.certificate_file):
        flash('Certificate file not found.')
        return redirect(url_for('admin_dashboard'))
    return send_file(achievement.certificate_file, as_attachment=True, download_name=f"certificate_{achievement.achievement_id}.pdf")

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        admin = Admin.query.filter_by(email=email).first()
        if admin and admin.password == password:
            session['admin_id'] = admin.admin_id
            session['user_type'] = 'admin'
            return redirect(url_for('admin_dashboard'))
        else:
            flash('Invalid email or password')
    return render_template('admin_login.html')

@app.route('/admin/dashboard')
def admin_dashboard():
    if 'admin_id' not in session or session.get('user_type') != 'admin':
        return redirect(url_for('admin_login'))
    department = request.args.get('department')
    achievement_type = request.args.get('type')
    year = request.args.get('year')
    month = request.args.get('month')
    faculty_name = request.args.get('faculty')
    query = db.session.query(Achievements, Faculty).join(Faculty, Achievements.faculty_id == Faculty.faculty_id)
    if department:
        query = query.filter(Faculty.department == department)
    if achievement_type:
        query = query.filter(Achievements.type == achievement_type)
    if year:
        query = query.filter(db.func.strftime('%Y', Achievements.date) == year)
    if month:
        query = query.filter(db.func.strftime('%Y-%m', Achievements.date) == month)
    if faculty_name:
        query = query.filter(Faculty.name == faculty_name)
    results = query.all()
    departments = db.session.query(Faculty.department).distinct().all()
    departments = [d[0] for d in departments]
    types = ['Publication', 'Conference', 'Workshop', 'Patent', 'Award']
    years = db.session.query(db.func.strftime('%Y', Achievements.date)).distinct().order_by(db.func.strftime('%Y', Achievements.date).desc()).all()
    years = [y[0] for y in years if y[0]]
    faculty_names = db.session.query(Faculty.name).distinct().all()
    faculty_names = [f[0] for f in faculty_names]
    
    # Analytics data
    dept_counts = db.session.query(Faculty.department, db.func.count(Achievements.achievement_id)).join(Achievements).filter(db.func.lower(Achievements.status) == 'approved').group_by(Faculty.department).all()
    dept_labels = [d[0] for d in dept_counts]
    dept_data = [d[1] for d in dept_counts]
    
    monthly_counts = db.session.query(db.func.strftime('%Y-%m', Achievements.date), db.func.count(Achievements.achievement_id)).filter(db.func.lower(Achievements.status) == 'approved').group_by(db.func.strftime('%Y-%m', Achievements.date)).order_by(db.func.strftime('%Y-%m', Achievements.date)).all()
    monthly_labels = [m[0] for m in monthly_counts]
    monthly_data = [m[1] for m in monthly_counts]
    
    type_counts = db.session.query(Achievements.type, db.func.count(Achievements.achievement_id)).filter(db.func.lower(Achievements.status) == 'approved').group_by(Achievements.type).all()
    type_labels = [t[0] for t in type_counts]
    type_data = [t[1] for t in type_counts]

    achievement_totals = {
        "publications": Achievements.query.filter_by(type='Publication').count(),
        "conferences": Achievements.query.filter_by(type='Conference').count(),
        "workshops": Achievements.query.filter_by(type='Workshop').count(),
        "patents": Achievements.query.filter_by(type='Patent').count(),
        "awards": Achievements.query.filter_by(type='Award').count()
    }
    
    return render_template('admin_dashboard.html', results=results, departments=departments, types=types, years=years, faculty_names=faculty_names, selected_dept=department, selected_type=achievement_type, selected_year=year, selected_month=month, selected_faculty=faculty_name,
                           dept_labels=json.dumps(dept_labels), dept_data=json.dumps(dept_data),
                           monthly_labels=json.dumps(monthly_labels), monthly_data=json.dumps(monthly_data),
                           type_labels=json.dumps(type_labels), type_data=json.dumps(type_data),
                           achievement_totals=achievement_totals)

@app.route('/admin/pending_achievements')
def pending_achievements():
    if 'admin_id' not in session or session.get('user_type') != 'admin':
        return redirect(url_for('admin_login'))
    pending_results = db.session.query(Achievements, Faculty).join(Faculty, Achievements.faculty_id == Faculty.faculty_id).filter(db.func.lower(Achievements.status) == 'pending').all()
    return render_template('pending_achievements.html', pending_results=pending_results)

@app.route('/admin/logout')
def admin_logout():
    session.pop('admin_id', None)
    session.pop('user_type', None)
    return redirect(url_for('home'))

@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

@app.route('/admin/add_faculty', methods=['GET', 'POST'])
def add_faculty():
    if 'admin_id' not in session or session.get('user_type') != 'admin':
        return redirect(url_for('admin_login'))
    if request.method == 'POST':
        name = request.form['name']
        email = request.form['email']
        password = request.form['password']
        department = request.form['department']
        designation = request.form['designation']
        existing = Faculty.query.filter_by(email=email).first()
        if existing:
            flash('Email already exists.')
            return redirect(url_for('add_faculty'))
        new_faculty = Faculty(name=name, email=email, password=password, department=department, designation=designation)
        db.session.add(new_faculty)
        db.session.commit()
        flash('Faculty added successfully.')
        return redirect(url_for('admin_dashboard'))
    return render_template('add_faculty.html')

@app.route('/admin/export_achievements')
def export_achievements():
    if 'admin_id' not in session or session.get('user_type') != 'admin':
        return redirect(url_for('admin_login'))
    month = request.args.get('month')
    format_type = request.args.get('format', 'csv')
    if not month:
        flash('Please select a month.')
        return redirect(url_for('admin_dashboard'))
    
    # Query achievements for the month
    start_date = datetime.datetime.strptime(month + '-01', '%Y-%m-%d').date()
    end_date = (start_date.replace(day=28) + datetime.timedelta(days=4)).replace(day=1) - datetime.timedelta(days=1)
    query = db.session.query(Achievements, Faculty).join(Faculty).filter(Achievements.date >= start_date, Achievements.date <= end_date, db.func.lower(Achievements.status) == 'approved')
    results = query.all()
    
    if format_type == 'csv':
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['Faculty Name', 'Department', 'Achievement Title', 'Type', 'Status', 'Date'])
        for achievement, faculty in results:
            writer.writerow([faculty.name, faculty.department, achievement.title, achievement.type, achievement.status, achievement.date])
        output.seek(0)
        return send_file(io.BytesIO(output.getvalue().encode('utf-8')), mimetype='text/csv', as_attachment=True, download_name=f'achievements_{month}.csv')
    elif format_type == 'excel':
        from openpyxl.styles import Font
        wb = Workbook()
        ws = wb.active
        ws.title = f'Achievements_{month}'
        
        # Headers with bold font
        headers = ['Faculty Name', 'Department', 'Achievement Title', 'Type', 'Status', 'Date']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
        
        # Data rows
        for row_num, (achievement, faculty) in enumerate(results, 2):
            ws.cell(row=row_num, column=1, value=faculty.name)
            ws.cell(row=row_num, column=2, value=faculty.department)
            ws.cell(row=row_num, column=3, value=achievement.title)
            ws.cell(row=row_num, column=4, value=achievement.type)
            ws.cell(row=row_num, column=5, value=achievement.status)
            ws.cell(row=row_num, column=6, value=str(achievement.date))
        
        # Auto-adjust column width
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'achievements_{month}.xlsx')
    elif format_type == 'pdf':
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font('Arial', 'B', 16)
        pdf.cell(0, 10, 'Faculty Achievement Report - VNRVJIET', ln=True, align='C')
        pdf.ln(5)
        
        pdf.set_font('Arial', 'B', 10)
        pdf.set_fill_color(102, 126, 234)
        pdf.set_text_color(255, 255, 255)
        
        # Table headers
        col_width = 38
        pdf.cell(col_width, 7, 'Faculty Name', border=1, fill=True)
        pdf.cell(col_width, 7, 'Department', border=1, fill=True)
        pdf.cell(col_width, 7, 'Achievement', border=1, fill=True)
        pdf.cell(col_width, 7, 'Type', border=1, fill=True)
        pdf.cell(col_width, 7, 'Date', border=1, fill=True)
        pdf.ln()
        
        pdf.set_font('Arial', '', 9)
        pdf.set_text_color(0, 0, 0)
        
        # Data rows
        for achievement, faculty in results:
            pdf.cell(col_width, 7, faculty.name[:20], border=1)
            pdf.cell(col_width, 7, faculty.department[:15], border=1)
            pdf.cell(col_width, 7, achievement.title[:20], border=1)
            pdf.cell(col_width, 7, achievement.type[:15], border=1)
            pdf.cell(col_width, 7, str(achievement.date), border=1)
            pdf.ln()
        
        pdf_output = io.BytesIO()
        pdf.output(pdf_output)
        pdf_output.seek(0)
        return send_file(pdf_output, mimetype='application/pdf', as_attachment=True, download_name=f'achievements_{month}.pdf')
    else:
        flash('Invalid format.')
        return redirect(url_for('admin_dashboard'))

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port)
